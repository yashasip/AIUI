import csv
import openpyxl


class DataRecordFile:
    def __init__(self, filePath) -> None:
        '''Handles Excel and Csv File data and related file operations
        Takes filePath: Path to the Excel and CSV File'''
        self.filePath = filePath
        self.fileType = self.getFileType() # determines extension of file
        self.sheetNames = [] # empty list to store sheet names of excel
        self.headersPresent = False # headers existence
        self.setupFile()

    def getFileType(self):
        for index, character in enumerate(self.filePath[::-1]):
            if character == ".":
                return self.filePath[-index:].lower()

    def setupFile(self): # setups excel and csv file using appropriate helper methods
        if self.fileType in ["xlsx", "xlsm", "xlsb", "xls"]:
            self.setupXL()
        else:
            self.setupCSV()

    def setupCSV(self): # helper function to setup csv file
        self.rawFile = open(self.filePath)
        self.file = csv.DictReader(self.rawFile)
        self.headers = self.setupHeaders(self.file.fieldnames)

    def setupXL(self): # helper function to setup excel file
        self.file = openpyxl.load_workbook(self.filePath)
        self.sheetNames = self.file.sheetnames
        self.currentSheet = self.file[self.sheetNames[0]]
        self.headers = self.setupHeaders(self.getHeadersXL())

    def setupHeaders(self, headerRow): # sets up headers, if file data with no headers is extracted new header values are generated and assigned
        for header in headerRow:
            if self.isNumber(header): # checks for existence of headers names, if all cells have numbers its not a header row
                return [f"Header {chr(65+i)}" for i in range(len(headerRow))]
        self.headersPresent = True 
        return headerRow

    def getHeadersXL(self): # get headers of each sheet for excel
        return [
            self.currentSheet.cell(row=1, column=column).value
            for column in range(1, self.currentSheet.max_column + 1)
        ]

    def changeSheetXL(self, changedSheet): # on sheet change
        self.currentSheet = self.file[changedSheet]
        self.headers = self.setupHeaders(self.getHeadersXL())  # check if required

    @staticmethod
    def isNumber(string): # checks if string is pure integer or float, returns True else False
        try:
            number = float(string)
        except ValueError:
            return False

        return True

    def cleanCsvData(self, selectedIndexes): # extracts only selected header values
        data = []
        self.rawFile.seek(0)  # set file pointer to first line
        file = csv.DictReader(  # new dictreader creation to extract data
            self.rawFile, range(len(self.headers))
        )
        if self.headersPresent:  # when headers present don't extract headers
            next(file)

        for row in file:
            rowData = []
            for item in selectedIndexes:
                rowData.append(float(row[item]))

            data += [rowData]
        return data

    def cleanXLData(self, selectedIndexes): # extracts only selected header values
        data = []
        for row in range(1, self.currentSheet.max_row + 1):
            if row == 1 and self.headersPresent:  # skip header if present
                continue
            rowData = []
            for column in range(1, self.currentSheet.max_column + 1):
                if column - 1 in selectedIndexes:
                    rowData.append(float(self.currentSheet.cell(row, column).value))
            data.append(rowData)

        return data

    def saveDataRecordFile(self, savePath, data): # takes path as string and data as 2d list, uses helper method to save files based on file extension
        if self.fileType == "csv":
            self.saveCsv(
                savePath, data
            )
        else:
            self.saveXL(
                savePath, data
            )

    def saveCsv(self, savePath, data): # helper method to save csv saves data from 2d list given as data
        outputFile = open(savePath, 'w',newline='') # opens a new file
        writer = csv.writer(outputFile)
        for row in data:
            writer.writerow(row)
        outputFile.close()

    def saveXL(self, savePath, data): # helper method to save excel saves data from 2d list given as data
        wb = openpyxl.Workbook()
        sheet = wb.active

        for row, rowData in enumerate(data):
            for column, cellData in enumerate(rowData):
                sheet.cell(row=row+1, column=column+1).value = cellData 

        wb.save(savePath)